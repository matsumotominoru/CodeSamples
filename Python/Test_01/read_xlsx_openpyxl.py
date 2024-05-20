#Get data from all sheets and output to a file.
import openpyxl
import csv
from datetime import datetime
from datetime import date

attriNameColumn = 3
now = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
wb = openpyxl.load_workbook(filename = 'C:\\Users\\999000_0F2582\\PythonTools\\テーブル定義書(WEB独自).xlsx')
file_name = "web_db_new_table_def_web_" + now + ".txt"
with open(file_name, 'w', encoding='utf-8') as f:

    for ws in wb:        
        if(ws["AA2"].value is not None):
            # テーブル名とテーブルID
            f.write(ws["AA2"].value + ':' + ws["Y2"].value + '\n')
            # テーブル項目
            for row in range(9, ws.max_row + 1):
                if(ws.cell(row,attriNameColumn).value is not None):
                    attriStr = ws.cell(row,attriNameColumn).value
                    f.write('\t' + attriStr + ':') 
                    f.write(ws.cell(row,attriNameColumn + 1).value.lower() + '\n')

    now2 = datetime.now().strftime("%d-%m-%Y-%H-%M-%S")
